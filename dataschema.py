import json
import logging
import os

import click
import openpyxl

from utils.compiler import DDLCompiler, DotCompiler
from utils.dot import make_img
from utils.schema import get_json_data, get_meta


@click.command()
@click.pass_context
@click.option(
    "--loglevel",
    "-l",
    type=click.Choice(["debug", "info", "warning", "error"]),
    default="info",
)
@click.argument("schema-jsons", nargs=-1)
@click.argument("output")
@click.option("--sql-dialect", "-d", default="mssql")
def main(ctx, loglevel, schema_jsons, output, sql_dialect="mssql"):
    """Script entry point."""
    if isinstance(loglevel, str):  # e.g. 'debug'/'DEBUG' -> logging.DEBUG
        loglevel = getattr(logging, loglevel.upper())
    ctx.ensure_object(dict)

    output = os.path.abspath(output)
    os.makedirs(os.path.dirname(output), exist_ok=True)
    suffix = output.split(".")[-1]

    json_data = {"resources": []}
    for schema_json in schema_jsons:
        schema_json = os.path.abspath(schema_json)
        assert output != schema_json
        _json_data = get_json_data(schema_json)
        json_data["resources"] += _json_data["resources"]

    if suffix == "json":
        with open(output, "w", encoding="utf-8") as file:
            json.dump(json_data, file, indent=2, ensure_ascii=False)
    elif suffix == "png":
        meta = get_meta(json_data, url=f"{sql_dialect}://")
        dot = str(DotCompiler().compile_tables(meta))
        with open(output, "wb") as file:
            file.write(make_img(dot))
    elif suffix == "sql":
        meta = get_meta(json_data, url=f"{sql_dialect}://")
        sql = str(DDLCompiler(meta.bind.url).compile_tables(meta))
        with open(output, "w", encoding="utf-8") as file:
            file.write(sql)
    elif suffix == "xlsx":
        wb = openpyxl.Workbook()
        for sname in wb.sheetnames:
            del wb[sname]
        for resource in json_data["resources"]:
            sht = wb.create_sheet(resource["name"])
            sht.cell(1, 1).value = "column_name"
            sht.cell(1, 2).value = "type"
            sht.cell(1, 3).value = "nullable"
            sht.cell(1, 4).value = "description"
            for i, field in enumerate(resource["schema"]["fields"]):
                sht.cell(i + 2, 1).value = field["name"]
                sht.cell(i + 2, 2).value = field["type"]
                sht.cell(i + 2, 3).value = (
                    "NULL" if field.get("nullable", True) else "NOT NULL"
                )
                sht.cell(i + 2, 4).value = field.get("description")
        wb.save(output)
    else:
        raise NotImplementedError(suffix)


if __name__ == "__main__":
    main()
